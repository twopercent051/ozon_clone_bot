from typing import List

from openpyxl.reader.excel import load_workbook
from pydantic import BaseModel

from create_bot import logger


# class ExcelItem(BaseModel):
#     ozon_id: int | str
#     # ozon_url: str
#     article: str


async def xlsx_parser(file: str) -> List[dict]:
    wb = load_workbook(filename=file)
    sh = wb.active
    result = []
    for row in sh.iter_rows(min_row=2):
        try:
            # ozon_id = row[0].value.split("/")[4].split("-")[-1]
            # ozon_id = row[0].value.split("/")[4]
            ozon_id = row[0].value
            article = row[1].value
            item = dict(ozon_id=ozon_id, article=article)
            if ozon_id and article:
                result.append(item)
        except Exception as ex:
            logger.warning(ex)
            result.append(None)
    return result
